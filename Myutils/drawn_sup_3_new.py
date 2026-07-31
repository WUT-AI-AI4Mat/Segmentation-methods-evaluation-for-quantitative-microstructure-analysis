import os
import random
import argparse
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import spearmanr
import warnings
import matplotlib.pyplot as plt
import seaborn as sns

warnings.filterwarnings('ignore')



































KEY_METRICS = ['miou', 'dice', 'precision', 'recall', 'acc', 'hd95', 'mae', 'nsd']


NEGATIVE_METRICS = ['hd95', 'mae']



def read_excel_data(file_path):
    """
    使用load_workbook读取Excel文件并返回DataFrame
    """
    print(f"正在读取文件: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    all_data = []
    for row in ws.iter_rows(values_only=True):
        all_data.append(list(row))

    wb.close()


    column_names = all_data[0]
    data_rows = all_data[1:]


    if len(data_rows) > 1:

        data_rows = data_rows[:]


    df = pd.DataFrame(data_rows, columns=column_names)


    df.columns = [str(col).strip().lower() for col in df.columns]


    for col in df.columns:
        if col not in ['方法名称', 'method', 'model']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    print(f"  读取完成: {len(df)} 行数据, 列名: {df.columns.tolist()}")
    return df


def process_negative_metrics(df):
    """
    对hd95、mae等指标取负数
    """
    df_processed = df.copy()

    for col in df_processed.columns:
        col_lower = col.lower().strip()
        if col_lower in NEGATIVE_METRICS:
            df_processed[col] = -df_processed[col].abs()
            print(f"  指标 '{col}' 已取负数")
    return df_processed


def extract_method_name(file_name):
    """
    从文件名提取模型名称（不区分大小写）
    注意：文件名对应的是模型名称，但需要与参考排名中的名称匹配
    """
    name = os.path.splitext(file_name)[0]
    print(name)
    return name


def remove_random_rows(df, remove_ratio=0.1):
    """
    从DataFrame中随机去除一定比例的行（至少一行）
    """
    n_rows = len(df)
    if n_rows <= 1:
        return df, []

    n_remove = max(1, int(n_rows * remove_ratio))
    remove_indices = random.sample(range(n_rows), n_remove)

    df_remaining = df.drop(index=remove_indices).reset_index(drop=True)
    print(f"  去除 {n_remove} 行 (共 {n_rows} 行)")

    return df_remaining, remove_indices



def compute_metric_ranking(df, metrics):
    """
    根据给定的指标计算排名（每个指标独立排名）
    注意：某些指标是越高越好（如miou, dice等），某些是越低越好（如hd95）
    这里我们统一按数值从大到小排名（因为hd95已经取了负数）
    """
    rankings = {}
    for metric in metrics:
        if metric in df.columns:


            sorted_values = df[metric].sort_values(ascending=False)

            ranks = pd.Series(range(1, len(sorted_values) + 1), index=sorted_values.index)
            rankings[metric] = ranks

    return rankings


def compute_spearman_correlation(computed_rankings, reference_rank_list, metric):
    """
    计算某个指标的排名与参考排名的斯皮尔曼相关性
    """


    method_names = list(computed_rankings[metric].index)
    print(method_names)
    print(reference_rank_list)
    computed_ranks = list(computed_rankings[metric].values)


    rank_dict = {}
    for i, name in enumerate(method_names):
        rank_dict[name] = computed_ranks[i]


    reference_ranks = list(range(1, len(reference_rank_list) + 1))


    computed_for_ref = []
    for model in reference_rank_list:
        if model in rank_dict:
            computed_for_ref.append(rank_dict[model])
        else:

            computed_for_ref.append(np.mean(list(rank_dict.values())))


    corr, p_value = spearmanr(reference_ranks, computed_for_ref)
    return corr



def main(folder_path, output_file, n_iterations=100, remove_ratio=0.3, seed=42):





    random.seed(seed)

    excel_files = [f for f in os.listdir(folder_path)
                   if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]

    if not excel_files:
        print("错误: 未找到Excel文件!")
        return

    print(f"\n找到 {len(excel_files)} 个Excel文件:")
    for f in excel_files:
        print(f"  - {f}")


    all_data = {}
    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        df = read_excel_data(file_path)


        df = process_negative_metrics(df)

        model_name = extract_method_name(file_name)
        all_data[model_name] = df


    results = []

    print(f"\n{'=' * 60}")
    print(f"开始执行 {n_iterations} 次随机去除和相关性计算")
    print(f"{'=' * 60}")

    for iteration in range(1, n_iterations + 1):
        print(f"\n迭代 {iteration}/{n_iterations}")


        iteration_seed = random.randint(1, 1000000)
        random.seed(iteration_seed)


        iteration_data = {}
        for model_name, df in all_data.items():
            df_remaining, _ = remove_random_rows(df, remove_ratio=remove_ratio)
            iteration_data[model_name] = df_remaining






        metric_means = {metric: [] for metric in KEY_METRICS}
        model_names_list = []


        for model_name, df in iteration_data.items():
            model_names_list.append(model_name)
            for metric in KEY_METRICS:
                if metric in df.columns:

                    mean_value = df[metric].mean()
                    metric_means[metric].append(mean_value)
                else:
                    metric_means[metric].append(np.nan)


        metric_df = pd.DataFrame(metric_means, index=model_names_list)


        rankings = {}
        for metric in KEY_METRICS:
            if metric in metric_df.columns:

                sorted_series = metric_df[metric].sort_values(ascending=False)
                ranks = pd.Series(range(1, len(sorted_series) + 1), index=sorted_series.index)
                rankings[metric] = ranks




        reference_column_name = 'last_column'

        dynamic_ranking = {}
        dynamic_mean_values = {}

        for model_name, df in iteration_data.items():
            mean_val = df.iloc[:,-1].mean()
            dynamic_mean_values[model_name] = mean_val











        sorted_mean_values = sorted(dynamic_mean_values.items(), key=lambda item: item[1], reverse=False)
        for rank, (model_name, mean_val) in enumerate(sorted_mean_values, start=1):
            dynamic_ranking[model_name] = rank

        print(f"  当前迭代的动态参考排名为: {dynamic_ranking}")



        iteration_result = {'iteration': iteration}
        for metric in KEY_METRICS:
            if metric in rankings:
                corr = compute_spearman_correlation(rankings, dynamic_ranking, metric)
                iteration_result[metric] = corr
                print(f"  指标 {metric}: 斯皮尔曼相关性 = {corr:.4f}")
            else:
                iteration_result[metric] = np.nan
                print(f"  指标 {metric}: 未找到，设为NaN")

        results.append(iteration_result)


    result_df = pd.DataFrame(results)


    columns_order = ['iteration'] + KEY_METRICS
    result_df = result_df[columns_order]


    result_df.to_excel(output_file, index=False)

    print(f"\n{'=' * 60}")
    print(f"计算完成！结果已保存到: {output_file}")
    print(f"{'=' * 60}")


    print("\n各指标斯皮尔曼相关性统计:")
    print(result_df[KEY_METRICS].describe().round(4))

    return result_df


def plot_spearman_boxplot(result_df, output_path='spearman_boxplot.png'):
    """
    绘制斯皮尔曼相关性的箱线图

    参数:
    - result_df: 包含斯皮尔曼相关性结果的DataFrame
    - output_path: 输出图片路径
    """

    plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False


    metrics_order = ['acc', 'precision', 'recall', 'dice', 'miou', 'hd95', 'nsd', 'mae']


    metric_labels = {
        'acc': 'Accuracy',
        'precision': 'Precision',
        'recall': 'Recall',
        'dice': 'Dice',
        'miou': 'mIoU',
        'hd95': 'HD95',
        'nsd': 'NSD',
        'mae': 'CE'
    }


    available_metrics = [m for m in metrics_order if m in result_df.columns]

    if len(available_metrics) < 2:
        print("警告：可用的指标数量不足，无法绘制箱线图")
        return


    plt.figure(figsize=(18, 9))


    sns.set_style("whitegrid")
    sns.set_palette("Set2")


    plot_data = []
    for metric in available_metrics:
        if metric in result_df.columns:
            values = result_df[metric].dropna().values
            for val in values:
                plot_data.append({
                    'Metric': metric_labels.get(metric, metric),
                    'Spearman Correlation': val
                })

    plot_df = pd.DataFrame(plot_data)


    ax = sns.boxplot(
        x='Metric',
        y='Spearman Correlation',
        data=plot_df,
        order=[metric_labels.get(m, m) for m in available_metrics],
        palette='Set2',
        width=0.6,
        showmeans=True,
        meanprops={
            'marker': 'D',
            'markerfacecolor': 'red',
            'markeredgecolor': 'red',
            'markersize': 6
        }
        ,
        flierprops={
            'marker': 'o',
            'markerfacecolor': 'orange',
            'markersize': 4,
            'alpha': 0.6
        }
    )















    plt.xlabel('Metric', fontsize=24, labelpad=10)
    plt.ylabel('Spearman Correlation Coefficient', fontsize=24, labelpad=10)


    plt.ylim(-1.1, 1.1)


    plt.axhline(y=0, color='gray', linestyle='--', alpha=0.3, linewidth=1)
    plt.axhline(y=0.5, color='green', linestyle=':', alpha=0.2, linewidth=1)
    plt.axhline(y=-0.5, color='red', linestyle=':', alpha=0.2, linewidth=1)


    plt.grid(axis='y', linestyle='--', alpha=0.3)


    plt.xticks(rotation=0, fontsize=24)
    plt.yticks(fontsize=20)


    plt.tight_layout()


    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"箱线图已保存至: {output_path}")


    plt.show()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run ranking robustness analysis.")
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output-xlsx", default="Supp_3a.xlsx")
    parser.add_argument("--output-figure", default="Supp_3a.pdf")
    parser.add_argument("--iterations", type=int, default=100)
    parser.add_argument("--remove-ratio", type=float, default=0.3)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()
    result_df = main(
        args.input_dir,
        args.output_xlsx,
        n_iterations=args.iterations,
        remove_ratio=args.remove_ratio,
        seed=args.seed,
    )
    plot_spearman_boxplot(result_df, output_path=args.output_figure)
