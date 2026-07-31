import pandas as pd
import numpy as np
import argparse
from openpyxl import load_workbook
from scipy.stats import spearmanr
from itertools import combinations
import matplotlib.pyplot as plt
import seaborn as sns


def plot_spearman_boxplot(result_df, output_path='spearman_boxplot.png'):
    """
    绘制斯皮尔曼相关性的箱线图

    参数:
    - result_df: 包含斯皮尔曼相关性结果的DataFrame
    - output_path: 输出图片路径
    """

    plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False


    metrics_order = ['acc', 'precision', 'recall', 'dice', 'miou', 'hd95', 'nsd', 'mae']


    metric_labels = {
        'acc': 'Accuracy',
        'precision': 'Precision',
        'recall': 'Recall',
        'dice': 'Dice',
        'miou': 'mIoU',
        'hd95': 'HD95',
        'nsd': 'NSD',
        'mae': 'CE'
    }


    available_metrics = [m for m in metrics_order if m in result_df.columns]

    if len(available_metrics) < 2:
        print("警告：可用的指标数量不足，无法绘制箱线图")
        return


    plt.figure(figsize=(18, 9))


    sns.set_style("whitegrid")
    sns.set_palette("Set2")


    plot_data = []
    for metric in available_metrics:
        if metric in result_df.columns:
            values = result_df[metric].dropna().values
            for val in values:
                plot_data.append({
                    'Metric': metric_labels.get(metric, metric),
                    'Spearman Correlation': val
                })

    plot_df = pd.DataFrame(plot_data)


    ax = sns.boxplot(
        x='Metric',
        y='Spearman Correlation',
        data=plot_df,
        order=[metric_labels.get(m, m) for m in available_metrics],
        palette='Set2',
        width=0.6,
        showmeans=True,
        meanprops={
            'marker': 'D',
            'markerfacecolor': 'red',
            'markeredgecolor': 'red',
            'markersize': 6
        }
        ,
        flierprops={
            'marker': 'o',
            'markerfacecolor': 'orange',
            'markersize': 4,
            'alpha': 0.6
        }
    )















    plt.xlabel('Metric', fontsize=24, labelpad=10)
    plt.ylabel('Spearman Correlation Coefficient', fontsize=24, labelpad=10)


    plt.ylim(-1.1, 1.1)


    plt.axhline(y=0, color='gray', linestyle='--', alpha=0.3, linewidth=1)
    plt.axhline(y=0.5, color='green', linestyle=':', alpha=0.2, linewidth=1)
    plt.axhline(y=-0.5, color='red', linestyle=':', alpha=0.2, linewidth=1)


    plt.grid(axis='y', linestyle='--', alpha=0.3)


    plt.xticks(rotation=0, fontsize=24)
    plt.yticks(fontsize=20)


    plt.tight_layout()


    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"箱线图已保存至: {output_path}")


    plt.show()
















def read_excel_data(file_path):
    """
    使用openpyxl读取Excel文件
    """
    print(f"正在读取文件: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active


    all_data = []
    for row in ws.iter_rows(values_only=True):
        all_data.append(list(row))

    wb.close()


    column_names = all_data[0]
    data_rows = all_data[1:]

    print(f"列名: {column_names}")
    print(f"数据行数: {len(data_rows)}")


    df = pd.DataFrame(data_rows, columns=column_names)
    return df


def preprocess_data(df):
    """
    数据预处理：
    1. 将第一列（方法名称）设置为索引
    2. 对hd95、mae和最后一列取负数
    """

    first_col = df.columns[0]
    print(f"第一列列名: {first_col}")


    df = df.set_index(first_col)


    cols = df.columns.tolist()


    target_cols_lower = ['hd95', 'mae']


    for col in cols:
        col_lower = col.lower().strip()
        if col_lower in target_cols_lower:
            print(f"对列 '{col}' 取负数")
            df[col] = pd.to_numeric(df[col], errors='coerce') * -1


    last_col = cols[-1]
    print(f"对最后一列 '{last_col}' 取负数")
    df[last_col] = pd.to_numeric(df[last_col], errors='coerce') * -1

    return df


def calculate_spearman_correlation(all_methods, methods_to_remove, df, metrics):
    """
    计算去掉特定方法后的斯皮尔曼相关性
    """

    remaining_methods = [m for m in all_methods if m not in methods_to_remove]

    if len(remaining_methods) < 2:
        return {metric: np.nan for metric in metrics}


    try:
        remaining_data = df.loc[remaining_methods]
    except KeyError as e:
        print(f"警告：方法 {e} 在数据中不存在")
        return {metric: np.nan for metric in metrics}


    last_col = df.columns[-1]
    reference_rank = remaining_data[last_col].rank(method='average').values

    results = {}
    for metric in metrics:
        if metric in df.columns:

            metric_values = pd.to_numeric(remaining_data[metric], errors='coerce')
            metric_rank = metric_values.rank(method='average').values

            if len(metric_rank) >= 2 and not np.all(np.isnan(metric_rank)):
                corr, _ = spearmanr(metric_rank, reference_rank)
                results[metric] = corr if not np.isnan(corr) else np.nan
            else:
                results[metric] = np.nan
        else:
            print(f"警告：指标 '{metric}' 在数据中不存在")
            results[metric] = np.nan

    return results


def main(file_path, output_path, figure_name):



    target_methods = [
        'DeepLabV3+',
        'SegFormer',
        'HQSAM+decoder+lora',
        'MatSAM+decoder+lora',
        'MicroSAM+decoder+lora',
        'SAM2+decoder+lora',
        'SwinUnet',
        'Unet'
    ]


    metrics = ['miou', 'dice', 'precision', 'recall', 'acc', 'hd95', 'mae', 'nsd']


    print("=" * 60)
    print("步骤1: 读取Excel文件")
    print("=" * 60)

    df = read_excel_data(file_path)

    print("\n步骤2: 数据预处理")
    print("=" * 60)
    df_processed = preprocess_data(df)

    print(f"\n预处理后的数据形状: {df_processed.shape}")
    print(f"索引（方法名称）: {df_processed.index.tolist()}")
    print(f"列名（指标）: {df_processed.columns.tolist()}")


    available_methods = set(df_processed.index.tolist())
    target_set = set(target_methods)
    missing_methods = target_set - available_methods

    if missing_methods:
        print(f"\n警告：以下方法未在数据中找到: {missing_methods}")
        print(f"数据中可用方法: {available_methods}")

        all_methods = [m for m in target_methods if m in available_methods]
    else:
        all_methods = target_methods

    print(f"\n参与计算的方法: {all_methods}")


    print("\n步骤3: 构建方法组合")
    print("=" * 60)

    all_combinations = []
    combination_labels = []


    for method in all_methods:
        all_combinations.append([method])
        combination_labels.append(method)


    if len(all_methods) >= 2:
        for combo in combinations(all_methods, 2):
            all_combinations.append(list(combo))
            combination_labels.append(' + '.join(combo))


    if len(all_methods) >= 3:
        for combo in combinations(all_methods, 3):
            all_combinations.append(list(combo))
            combination_labels.append(' + '.join(combo))

    print(f"总组合数: {len(all_combinations)}")
    if len(all_methods) >= 2:
        print(f"单方法组合: {len(all_methods)}个")
        print(f"两方法组合: {len(list(combinations(all_methods, 2)))}个")
    if len(all_methods) >= 3:
        print(f"三方法组合: {len(list(combinations(all_methods, 3)))}个")


    print("\n步骤4: 计算斯皮尔曼相关性")
    print("=" * 60)

    results_list = []

    for i, (combo, label) in enumerate(zip(all_combinations, combination_labels)):

        if (i + 1) % 10 == 0 or i == 0 or i == len(all_combinations) - 1:
            print(f"处理组合 {i + 1}/{len(all_combinations)}: {label}")


        corr_results = calculate_spearman_correlation(
            all_methods, combo, df_processed, metrics
        )


        result_row = {'组合名称': label}
        for metric in metrics:
            value = corr_results.get(metric, np.nan)
            if not np.isnan(value):
                result_row[metric] = round(value, 4)
            else:
                result_row[metric] = np.nan

        results_list.append(result_row)


    print("\n步骤5: 保存结果")
    print("=" * 60)


    columns_order = ['组合名称'] + metrics
    result_df = pd.DataFrame(results_list, columns=columns_order)


    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='斯皮尔曼相关性', index=False)

    print(f"结果已保存至: {output_path}")
    print(f"结果形状: {result_df.shape}")
    print(f"结果列名: {result_df.columns.tolist()}")


    print("\n前5行结果预览:")
    print(result_df.head(5).to_string())


    print("\n结果统计（仅针对数值列）:")

    for metric in metrics:
        valid_values = result_df[metric].dropna()
        if len(valid_values) > 0:
            print(
                f"{metric}: 均值={valid_values.mean():.4f}, 标准差={valid_values.std():.4f}, 有效数据数={len(valid_values)}")
        else:
            print(f"{metric}: 无有效数据")

    print("\n程序执行完成!")

    print("\n步骤6: 绘制斯皮尔曼相关性箱线图")
    print("=" * 60)

    plot_spearman_boxplot(result_df, figure_name)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyze metric ranking correlations.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-xlsx", default="Supp_2g.xlsx")
    parser.add_argument("--output-figure", default="Supp_2g.pdf")
    args = parser.parse_args()
    main(args.input, args.output_xlsx, args.output_figure)
