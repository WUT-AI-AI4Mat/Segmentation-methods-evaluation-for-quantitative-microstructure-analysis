import argparse
import glob
import os

import matplotlib.pyplot as plt
from openpyxl import load_workbook


COLORS = {
    "Unet": "#ff7f0e",
    "Deep": "#17becf",
    "SegFormer": "#bcbd22",
    "Swin": "#7f7f7f",
    "S2": "#1f77b4",
    "HQ": "#d62728",
    "uS": "#2ca02c",
    "Mat": "#9467bd",
    "S2-L": "#6baed6",
    "HQ-L": "#ff9896",
    "uS-L": "#98df8a",
    "Mat-L": "#c5b0d5",
}

FILE_COLORS = (
    ("Grain_hqsam_LoRA_Decoder", "HQ-L"),
    ("Grain_matsam_LoRA_Decoder", "Mat-L"),
    ("Grain_microsam_LoRA_Decoder", "uS-L"),
    ("Grain_SAM2_LoRA_Decoder", "S2-L"),
    ("Grain_DeepLabV3+", "Deep"),
    ("Grain_hqsam", "HQ"),
    ("Grain_matsam", "Mat"),
    ("Grain_microsam", "uS"),
    ("Grain_SAM2", "S2"),
    ("Grain_SegFormer", "SegFormer"),
    ("Grain_SwinUnet", "Swin"),
    ("Grain_Unet", "Unet"),
)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Plot CE against quantitative analysis error."
    )
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output", default="Fig_3d.pdf")
    return parser.parse_args()


def get_color(file_path):
    for marker, color_key in FILE_COLORS:
        if marker in file_path:
            return COLORS[color_key]
    return "#7f7f7f"


def load_points(file_path):
    workbook = load_workbook(file_path, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], []

    x_values = []
    y_values = []
    for row in rows[1:]:
        if row is None or len(row) < 11:
            continue
        x_value, y_value = row[10], row[-1]
        if x_value is None or y_value is None:
            continue
        try:
            x_values.append(float(x_value))
            y_values.append(float(y_value))
        except (TypeError, ValueError):
            continue
    return x_values, y_values


def main():
    args = parse_args()
    excel_files = glob.glob(os.path.join(args.input_dir, "*.xlsx"))
    if not excel_files:
        raise FileNotFoundError(f"No .xlsx files found in {args.input_dir}")

    plt.figure(figsize=(10, 7))
    plotted = 0
    for file_path in excel_files:
        try:
            x_values, y_values = load_points(file_path)
        except Exception as exc:
            print(f"Failed to process {os.path.basename(file_path)}: {exc}")
            continue

        if not x_values:
            print(f"Skipped {os.path.basename(file_path)}: no numeric data.")
            continue

        plt.scatter(
            x_values,
            y_values,
            c=get_color(file_path),
            alpha=0.4,
            label=os.path.basename(file_path),
        )
        plotted += 1

    if plotted == 0:
        raise ValueError("No valid numeric data was found in the input files.")

    plt.xlabel("CE", size=20)
    plt.ylabel("Quantitative Analysis Error (%)", size=20)
    plt.tick_params(length=5, pad=5, labelsize=16)
    plt.grid(True)
    plt.yscale("symlog")
    plt.xscale("symlog")
    plt.ylim(-0.5, 900)
    plt.xlim(-0.5, 500)
    plt.tight_layout()
    plt.savefig(args.output)
    plt.show()


if __name__ == "__main__":
    main()
